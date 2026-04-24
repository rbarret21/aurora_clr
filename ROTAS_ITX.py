
# -*- coding: utf-8 -*-
# FINAL ROTAS_ITX.py — GERADOR ITX (COM CSC NOVO + AGRUPAMENTO POR SOFTX)

import io, os, re
from datetime import datetime
from flask import Flask, render_template, request, send_file, send_from_directory, flash, redirect, url_for
import pandas as pd
from collections import defaultdict
from io import BytesIO
from openpyxl import load_workbook, Workbook
from copy import copy
from openpyxl.utils import get_column_letter
from collections import deque

LOG_BUFFER = deque(maxlen=500)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
GENERATED_DIR = os.path.join(BASE_DIR, "generated")
TXT_DIR = os.path.join(GENERATED_DIR, "txt")
XLSX_DIR = os.path.join(GENERATED_DIR, "xlsx")

os.makedirs(TXT_DIR, exist_ok=True)
os.makedirs(XLSX_DIR, exist_ok=True)


ALLOWED_EXTENSIONS = {'.csv', '.xls', '.xlsx'}
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MODELOS_DIR = os.path.join(BASE_DIR, 'modelos')
MODELO_ATIVACAO = os.path.join(MODELOS_DIR, "modelo_tab_ativ_rotas.xlsx")

# =========================
# IFMI MAP (mantido)
# =========================
IFMI_MAP = {
    '10.252.53.1':'1700','10.252.53.2':'1701','10.252.53.3':'1704','10.252.53.4':'1705',
    '10.6.222.42':'1700','10.6.222.43':'1702','10.6.222.44':'1704','10.6.222.45':'1705',
    '10.252.83.4':'1700','10.252.83.5':'1701','10.252.83.6':'1704','10.252.83.7':'1705',
    '10.0.0.36':'1705','10.252.20.50':'1700','10.252.20.51':'1701','10.252.20.52':'1704',
    '10.252.84.5':'1700','10.252.84.6':'1701','10.252.84.7':'1704','10.252.84.8':'1705',
    '10.6.222.18':'1700','10.6.222.19':'1702','10.6.222.20':'1704','10.6.222.21':'1705',
    '10.252.28.1':'1700','10.252.28.2':'1701','10.252.28.6':'1704','10.252.28.7':'1705',
    '10.252.18.1':'132','10.252.18.2':'133','10.252.18.4':'134','10.252.18.5':'135',
    '10.0.0.68':'132','10.252.21.50':'133','10.252.21.51':'134','10.252.21.52':'135',
    '10.252.19.1':'132','10.252.19.2':'133','10.252.19.4':'134','10.252.19.5':'135',
    '10.252.29.1':'132','10.252.29.2':'133','10.252.29.3':'134','10.252.29.4':'135',
    '10.252.66.1':'132','10.252.66.2':'133','10.252.66.3':'134','10.252.66.4':'135',
    '10.252.17.1':'132','10.252.17.2':'133','10.252.17.3':'134',
    '10.252.44.1':'132','10.252.44.2':'133','10.252.44.3':'134','10.252.44.4':'135',
    '10.252.27.1':'1700','10.252.27.2':'1701','10.252.27.3':'1702','10.252.27.6':'1703',
    '10.252.82.4':'1700','10.252.82.5':'1701','10.252.82.6':'1702','10.252.82.7':'1703'
}

SOFTX_SIGLA_MAP = {
    "BHE MGC ES": "HBHB",
    "CTA MGC MC": "HCTA",
    "FLA MGC PV": "HFLA",
    "SPO MGC IG": "HSPO",
    "SPO MGC IG2": "HSPC",
    "SPO MGC PH": "HSPA",
    "SPO MGC PH2": "HSPB",
    "SPO MGC MB": "HSMI",
    "SPO MGC LP": "HSLP",
    "RJO MGC AM": "HRJO",
    "RJO MGC AM2": "HRJB",
    "RJO MGC EN": "HRJA",
    "SDR MGC RC": "HDSA",
    "BSA MGC SU": "HBSA",
    "BRU MGC AL": "HBRA",
    "REC MGC AM": "HRCB",
}


VALID_TRAFS = {"SMP", "STFC", "IMT"}
VALID_TIPOS = {"LC", "LD", "MS", "TP"}


VALID_CN = set()

VALID_CN.update([11, 12, 13, 14, 15, 16, 17, 18, 19])          # 11–19
VALID_CN.update([21, 22, 24, 27, 28])
VALID_CN.update([31, 32, 33, 34, 35, 37, 38])          # 31–35
VALID_CN.update([41, 42, 43, 44, 45, 46, 47, 48, 49])          # 41–49
VALID_CN.update([51, 53, 54, 55])
VALID_CN.update([61, 62, 63, 64, 65, 66, 67, 68, 69])          # 61–69
VALID_CN.update([71, 73, 74, 75, 77, 79])
VALID_CN.update([81, 82, 83, 84, 85, 86, 87, 88, 89])          # 81–89
VALID_CN.update([91, 92, 93, 94, 95, 96, 97, 98, 99])          # 91–99


# =========================
# TEMPLATES ITX (mantidos)
# =========================
TPL_ADD_SRT = 'ADD SRT: SRC=TTTT, O=320, SRN="RRRR", RENT=URT;'
TPL_ADD_SIPTG = (
    'ADD SIPTG: TG=TTTT, CSC=CCCC, TGN="SSSS", SRT=TTTT, RCHS=0, OTCS=65535, HCIC=NNNN, LCIC=MMMM, '
    'ST=NGNN, NOAA=YES, TCN=NO, CCN=NO, CNA=DEFUNKN1, CCRN=NO, CRNA=DEFUNKN1, '
    'ICR=LCO-1&LC-1&LCT-1&NTT-1&ITT-1&INTT-1&IITT-1&IOLT-1&CCR1-1&CCR2-1&CCR3-1&CCR4-1&CCR5-1&CCR6-1&CCR7-1&CCR8-1&CCR9-1&CCR10-1&CCR11-1&CCR12-1&CCR13-1&CCR14-1&CCR15-1&CCR16-1, '
    'OCR=LCO-1&LC-1&LCT-1&NTT-1&ITT-1&INTT-1&IITT-1&IOLT-1&CCR1-1&CCR2-1&CCR3-1&CCR4-1&CCR5-1&CCR6-1&CCR7-1&CCR8-1&CCR9-1&CCR10-1&CCR11-1&CCR12-1&CCR13-1&CCR14-1&CCR15-1&CCR16-1, '
    'CAMA=NO, IT=NO, ABT=NO, ISBF=NO, UHB=NO, SFPARA=SVR17-1&SVR18-1, SGCTRL=SVR10-1, SGCTRLS=SVR25-1, '
    'EXTOLP=NO, RIVAL=165, STM=0, SRM=1, MST=SOALW, SST=SIALW, PL=PI, ACODEC=G729, VIDEOS=SUPPORT, CHBF=NO, '
    'MAX=65535, MAXO=65535, SI=1800, CODECS=PCMA-1&PCMU-1&G7231-1&G726-1&G728-1&G729-1&MPEG4A-1&S2833-1&G726_40-1&G726_32-1&G726_24-1&G726_16-1&H261-1&H263-1&MPEG4V-1&H264-1&AMR_WB-1&T120-1&T38-1&AMR-1&CLEARMODE-0&ILBC-1&SPEEX-1&G722-1&GSM_FR-1, '
    'SELMODE=DIST, SRO=YES, IRCMFLAG=NO, ORCMFLAG=NO;'
)
TPL_ADD_SIPIPPAIR = 'ADD SIPIPPAIR: TG=TTTT , IMN=FFFF, OSU="PPPP:5060", DH=No, LSRVP=5060;'
TPL_MOD_SIPTG_ITABT = 'MOD SIPTG: TG=TTTT, IT=YES, ABT=YES;'
TPL_MOD_BTG = 'MOD BTG: TG=TTTT, BLS=BLK, BLD=INOUT;'
TPL_MOD_SIPTG_UHB = 'MOD SIPTG: TG=TTTT, UHB=NORMAL, NHB=10, XHB=20,EA=YES, SGCTRLT=SVR7-1;'
TPL_ADD_TGDSG = 'ADD TGDSG: TG=TTTT, DSG=80XX;'

TPL_RT_PERC_HEAD = 'ADD RT:R=TTTT,IDTP=UNKNOWN,NAMECFG=NO,SNCM=SRT,SRST=PERC,PERCCFG=YES'
TPL_RT_SEQ_HEAD  = 'ADD RT: R=TTTT, IDTP=UNKNOWN, NAMECFG=NO, SNCM=SRT, SRST=SEQ'
TPL_RT_TAIL = ', STTP=INVALID, REM=NO;'
TPL_ADD_RTANA_AAAA = 'ADD RTANA: RAN="AAAA", RSC=OOVXX, RSSC=65534, TM=TMM, R=TTTT;'

# =========================
# CSC (SMP + STFC + CN 5X)
# =========================

def compute_cccc(tipo, cn_str, traf):
    tipo_u = str(tipo).strip().upper()
    traf_u = str(traf).strip().upper()

    cn = str(cn_str).zfill(2)
    x = cn[1]
    is5 = cn.startswith("5")
#    log(f"[DEBUG] compute_cccc -> tipo={tipo_u} traf={traf_u} cn={cn}")
    # =========================
    # TP SEMPRE FIXO
    # =========================
    if tipo_u == "TP":
        return "3720"

    # =========================
    # REGRA SMP (PRIORIDADE)
    # =========================
    if traf == "SMP":
        if tipo_u == "LC":
            return f"440{x}"
        elif tipo_u in ("LD", "MS"):
            return f"441{x}"
        # MS e outros continuam STFC

    # =========================
    # REGRA STFC (PADRÃO)
    # =========================
    if is5:
        if tipo_u == "LC":
            return f"462{x}"
        elif tipo_u in ("LD", "MS"):
            return f"463{x}"
        else:
            return f"462{x}"
    else:
        if tipo_u == "LC":
            return f"460{x}"
        elif tipo_u in ("LD", "MS"):
            return f"461{x}"
        else:
            return f"460{x}"
#    print(
#        f"DEBUG RESULT -> CSC={cccc}"
#    )


# ======== RESTANTE DAS FUNÇÕES MANTIDAS ========
def compute_v(traf, tipo):
    traf_u = str(traf).strip().upper()
    tipo_u = str(tipo).strip().upper()
    V_MAP_STFC = {'LC': '0', 'LD': '1', 'MS': '2', 'TP': '3'}
    V_MAP_SMP  = {'LC': '5', 'LD': '6', 'MS': '7', 'TP': '8'}
    return (V_MAP_SMP if traf_u=='SMP' else V_MAP_STFC).get(tipo_u, '4')


def norm_cn(v):
    try: return str(int(v)).zfill(2)
    except: return str(v).strip().zfill(2)


def allowed_file(fname):
    return os.path.splitext(fname)[1].lower() in ALLOWED_EXTENSIONS


def load_dataframe(fs):
    ext = os.path.splitext(fs.filename)[1].lower()
    if ext == '.csv':
        data = fs.stream.read()
        try: return pd.read_csv(BytesIO(data))
        except: return pd.read_csv(BytesIO(data), sep=';')
    elif ext in ('.xls','.xlsx'):
        return pd.read_excel(fs, engine='openpyxl' if ext=='.xlsx' else 'xlrd')
    raise ValueError('Formato não suportado.')


# SRSEQ LOGIC
import re

def parse_sr_index(val, fallback):
    try:
        m=re.search(r'SR(\d{1,2})', str(val).upper())
        if m:
            n=int(m.group(1))
            if 1<=n<=16: return n
    except: pass
    return fallback


def determine_mode_and_members(rows):
    mode = 'SEQ' if any(str(r['PART']).strip().upper()=='SEQ' for r in rows) else 'PERC'
    members=[]; nxt=1
    for r in rows:
        sr=parse_sr_index(r['SRSEQ'], nxt)
        nxt=max(nxt, sr+1)
        item={'idx': sr, 'tgn': r['TGN']}
        if mode=='PERC':
            try: p=int(float(str(r['PART']).strip())); p=max(0,min(100,p))
            except: p=50
            item['perc']=p
        members.append(item)
    members.sort(key=lambda x:x['idx'])
    if mode=='PERC' and len(members)==1:
        members[0]['perc']=100
    return mode, members

# =============================================================
#  GERAR TXT — AGRUPANDO POR SOFTX E CABEÇALHO EM UMA LINHA
# =============================================================

def gerar_txt(df):

    teve_erro_critico = False # MARCA ERRO NA PLANILHA DE ENTRADA

    df=df.copy(); df.columns=[str(c).strip() for c in df.columns]

    required=['CN','TIPO DE ROTA','NOME DA ROTA','TGN','SESSOES','SOFTX','IP SX',
              'SBC','IP SBC','SRSEQ','PART','PST','CODIGO DA OPERADORA','TRAF','DOC','PTI']
    for c in required:
        if c not in df.columns:
            raise ValueError(f'Coluna ausente: {c}')

    recs=[]
    
    for _, row in df.iterrows():
        
        cn = validar_cn(row["CN"])
        if cn is None:
            log("[ERROR] Linha ignorada por CN inválido")
            teve_erro_critico = True
            continue
        
        tgn = str(int(row["TGN"]))
        linha_info = f"(CN={cn}, TGN={tgn})"
        tgn = str(int(row["TGN"]))
    
        linha_info = f"(CN={cn}, TGN={tgn})"
    
        tipo = validar_campo(
            "TIPO DE ROTA",
            row["TIPO DE ROTA"],
            VALID_TIPOS,
            default=None,
            linha_info=linha_info
        )
        if tipo is None:
            log(f"[ERROR] {linha_info} Linha ignorada por TIPO inválido")
            teve_erro_critico = True
            continue
    
        traf = validar_campo(
            "TRAF",
            row["TRAF"],
            VALID_TRAFS,
            default="none",
            linha_info=linha_info
        )
        nome=str(row['NOME DA ROTA']).strip()
        tgn=str(int(row['TGN']))
        sessoes=int(row['SESSOES'])
        softx, sigla_softx = validar_softx(row["SOFTX"], linha_info)
        ip_sx=str(row['IP SX']).strip()
        sbc=str(row['SBC']).strip()
        ip_sbc=str(row['IP SBC']).strip()
        srseq=str(row['SRSEQ']).strip()
        part=str(row['PART']).strip()
        pst=str(row['PST']).strip()
        codop=str(int(row['CODIGO DA OPERADORA']))
        doc=str(row['DOC']).strip()
        pti=str(row['PTI']).strip()

        ssss=f'SIP_{pst}_{traf}_CN{cn}_{tipo}_{sbc}_{doc}'
        aaaa=f'{pst}_CN{cn}_{doc}'
        nnnn=str(sessoes+1)
        mmmm=str(sessoes)
        cccc=compute_cccc(tipo, cn, traf)
        ifmi=IFMI_MAP.get(ip_sx.replace(' ',''), '')
        v=compute_v(traf, tipo)
        oov=f"{codop}{v}{cn}"

        recs.append({
            'SOFTX':softx,'CN':cn,'TIPO':tipo,'NOME':nome,'TGN':tgn,'SESSOES':sessoes,
            'IP_SX':ip_sx,'SBC':sbc,'IP_SBC':ip_sbc,'SRSEQ':srseq,'PART':part,'PST':pst,
            'CODOP':codop,'TRAF':traf,'DOC':doc,'PTI':pti,
            'SSSS':ssss,'AAAA':aaaa,'NNNN':nnnn,'MMMM':mmmm,'CCCC':cccc,'IFMI':ifmi,'OOV':oov
        })

    if teve_erro_critico:
        log("[ERROR] Erros críticos encontrados. TXT não será gerado.")
        return None

    # =============================================================
    # AGRUPAR POR SOFTX
    # =============================================================
    soft_groups=defaultdict(list)
    for r in recs:
        soft_groups[r['SOFTX']].append(r)

    out=[]
    for soft,rows in soft_groups.items():
        out.append(f"// ============= SOFTX: {soft} =============\n//")

        # Agora grupo CN+TIPO
        cn_groups=defaultdict(list)
        for r in rows:
            cn_groups[(r['CN'], r['TIPO'])].append(r)

        for key, group in cn_groups.items():
            # Cabeçalho único
            r=group[0]
            head=f"// CN {r['CN']} | TIPO {r['TIPO']} | ROTA {r['NOME']} | TGN {r['TGN']} | SBC {r['SBC']} | SOFTX {r['SOFTX']} ({r['IP_SX']}) | PTI {r['PTI']}"
            out.append(head)

            for r in group:
                TTTT=r['TGN']; RRRR=r['NOME']; SSSS=r['SSSS']; PPPP=r['IP_SBC']; FFFF=r['IFMI']
                NNNN=r['NNNN']; MMMM=r['MMMM']; CCCC=r['CCCC']

                def sub(x):
                    return (x.replace('TTTT',TTTT).replace('RRRR',RRRR).replace('SSSS',SSSS)
                              .replace('PPPP',PPPP).replace('FFFF',FFFF)
                              .replace('NNNN',NNNN).replace('MMMM',MMMM)
                              .replace('CCCC', str(CCCC)))

                out.append(sub(TPL_ADD_SRT))
                out.append(sub(TPL_ADD_SIPTG))
                out.append(sub(TPL_ADD_SIPIPPAIR))
                out.append(sub(TPL_MOD_SIPTG_ITABT))
                out.append(sub(TPL_MOD_BTG))
                out.append(sub(TPL_MOD_SIPTG_UHB))
                if r['TIPO'] in ('LC', 'MS'): out.append(sub(TPL_ADD_TGDSG))
                out.append('//')

            # bloco RT/RTANA
            mode, members=determine_mode_and_members(group)
            sr_parts=''.join([f", SR{m['idx']}={m['tgn']}" for m in members])
            if mode=='SEQ':
                rt_line=TPL_RT_SEQ_HEAD.replace('TTTT',members[0]['tgn'])+sr_parts+TPL_RT_TAIL
            else:
                perc_parts=''.join([f", PERC{m['idx']}={m.get('perc',0)}" for m in members])
                rt_line=TPL_RT_PERC_HEAD.replace('TTTT',members[0]['tgn'])+sr_parts+perc_parts+TPL_RT_TAIL

            r0 = group[0]
            out.append(f"// RT/RTANA PARA CN {r0['CN']} | TIPO {r0['TIPO']} | MODE {mode}")
            out.append(rt_line)
            
            nova_aaaa = f"{r0['PST']}_{r0['TIPO']}_CN{r0['CN']}_{r0['DOC']}"
            
            rtana = (
                 TPL_ADD_RTANA_AAAA
                .replace('TTTT', r0['TGN'])
                .replace('AAAA', nova_aaaa)
                .replace('OOVXX', r0['OOV'])
            )
             
            out.append(rtana)
            out.append('//')
            
    return "\n".join(out) + "\n"
    
# =====================================================
# XLSX – PLANILHA DE ATIVAÇÃO
# =====================================================
def gerar_planilha_ativacao(df):

    teve_erro_critico = False
    
    wb_modelo = load_workbook(MODELO_ATIVACAO)
    ws_modelo = wb_modelo.active

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    hoje_br = datetime.now().strftime("%d/%m/%Y")

    
    for _, r in df.iterrows():
    
        #  CN (primeiro de tudo)
        cn = validar_cn(r["CN"])
        if cn is None:
            log("[ERROR] Linha ignorada por CN inválido")
            teve_erro_critico = True
            continue
    
        # TGN (antes do linha_info)
        try:
            tgn = str(int(r["TGN"]))
        except Exception:
            log(f"[ERROR] (CN={cn}) TGN inválido: '{r['TGN']}'")
            teve_erro_critico = True
            continue
    
        # linha_info 
        linha_info = f"(CN={cn}, TGN={tgn})"
    
        # TIPO
        tipo = validar_campo(
            "TIPO DE ROTA",
            r["TIPO DE ROTA"],
            VALID_TIPOS,
            default=None,
            linha_info=linha_info
        )
        if tipo is None:
            log(f"[ERROR] {linha_info} Linha ignorada por TIPO inválido")
            teve_erro_critico = True
            continue
        
        if tipo == "MS":
            tipo_display = "MISTO LC/LD"
        else:
            tipo_display = tipo

    
        # TRAF
        traf = validar_campo(
            "TRAF",
            r["TRAF"],
            VALID_TRAFS,
            default="STFC",
            linha_info=linha_info
        )
    
        # SOFTX
        softx, sigla_softx = validar_softx(r["SOFTX"], linha_info)

        pti = str(r["PTI"]).strip()
        pppp = str(r["IP SBC"]).strip()
        sigla_softx = SOFTX_SIGLA_MAP.get(softx, "")
        rtp = "SIM" if tipo == "TP" else "NÃO"
        nome_aba = f"{tipo}_{cn}_{pti}"[:31]
        ws = wb_out.create_sheet(nome_aba)
        accip = str(r.get("IP SIN SBC ACC", "")).strip()
        
        for col_letter, dim in ws_modelo.column_dimensions.items():
            ws.column_dimensions[col_letter].width = dim.width

        for i, row_modelo in enumerate(ws_modelo.iter_rows(), start=1):
            for j, cell_modelo in enumerate(row_modelo, start=1):
                valor = cell_modelo.value
                novo_valor = valor
                if isinstance(valor, str):
                    novo_valor = (
                        valor.replace("CNNUMBER", cn)
                        .replace("TTTT", str(r["TGN"]))
                        .replace("RRRR", str(r["NOME DA ROTA"]))
                        .replace("SESSOES", str(r["SESSOES"]))
                        .replace("SOFTX", str(r["SOFTX"]))
                        .replace("PPPP", str(r["IP SBC"]))
                        .replace("IP SX", str(r["IP SX"]))
                        .replace("BBBB", sigla_softx)
                        .replace("TYPE", tipo_display)
                        .replace("RTP", rtp)
                        .replace("SBCNAME", str(r["SBC"]))
                        .replace("TRAF", str(r["TRAF"]))
                        .replace("PST", str(r["PST"]))
                        .replace("DOC", str(r["DOC"]))
                        .replace("ACCIP", accip)
                        .replace("PTINUMBER", str(r["PTI"]))
                        .replace("ROTA EBT", str(r.get("ROTA EBT", "")))
                        .replace("IP SIN SBC ACC", str(r.get("IP SIN SBC ACC", "")))
                        .replace("IP OPERADORA", str(r.get("IP OPERADORA", "")))
                        .replace("DD/MM/AAAA", hoje_br)
                    )
                nova = ws.cell(row=i, column=j, value=novo_valor)
                                
                if cell_modelo.has_style:
                    nova.font = copy(cell_modelo.font)
                    nova.border = copy(cell_modelo.border)
                    nova.fill = copy(cell_modelo.fill)
                    nova.number_format = cell_modelo.number_format
                    nova.alignment = copy(cell_modelo.alignment)
                    nova.protection = copy(cell_modelo.protection)

    if teve_erro_critico:
        log("[ERROR] Erros críticos encontrados. Planilha não será gerada.")
        return None



    bio = BytesIO()
    wb_out.save(bio)
    bio.seek(0)
    return bio
    

def safe_filename(value):
    """
    Converte valor em string segura para nome de arquivo
    """
    return re.sub(r'[\\/:*?"<>|]+', "_", str(value)).strip()



# =========================
# CONVERTE ABA EXCEL → HTML
# =========================
def aba_para_html(ws):
    html = []
    html.append("""
    <html>
    <body style="font-family:Calibri, Arial; font-size:11pt">
    <table border="1" cellspacing="0" cellpadding="6"
           style="border-collapse:collapse; width:100%">
    """)

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        html.append("<tr>")

        for col_idx, cell in enumerate(row, start=1):
            value = "" if cell.value is None else str(cell.value)

            # CABEÇALHO
            if row_idx == 1:
                html.append(
                    f'<td style="background-color:#FFA500; '
                    f'font-weight:bold; text-align:center;">{value}</td>'
                )

            else:
                # 🔹 Coluna B → alinhamento à direita
                if col_idx == 2:
                    html.append(
                        f'<td style="text-align:right;">{value}</td>'
                    )
                else:
                    html.append(f"<td>{value}</td>")

        html.append("</tr>")

    html.append("""
    </table>
    </body>
    </html>
    """)

    return "".join(html)


# =============================================================
# LOOP ENVIAR E-MAIL
# =============================================================


import socket
import smtplib

from email.message import EmailMessage

def enviar_email_smtp_html(destinatarios, assunto, html_body, contexto=""):
    try:
        msg = EmailMessage()
        msg["From"] = "robinson.barreto@gmail.com"
        msg["To"] = destinatarios
        msg["Subject"] = assunto
        msg.set_content("Este e-mail contém HTML.")
        msg.add_alternative(html_body, subtype="html")

        with smtplib.SMTP("smtp.gmail.com", 587, timeout=60) as s:
            s.ehlo()
            s.starttls()
            s.login("robinson.barreto@gmail.com", "nbtm pxol rjch pgds")
            s.send_message(msg)

        return True  # ✅ enviado com sucesso
    except socket.timeout:
        log(f"[ERROR] Falha ao enviar e-mail: conexão com o servidor expirou (timeout). {contexto}")
        return False

    except smtplib.SMTPException as e:
        log(f"[ERROR] Falha ao enviar e-mail: erro SMTP ({e}). {contexto}")
        return False

    except Exception as e:
        log(f"[ERROR] Falha inesperada ao enviar e-mail: {e}. {contexto}")
        return False

        
# =============================================================
# JANELA DE LOGS
# ============================================================= 
 
 # Buffer de logs (últimas N linhas)

def log(msg):
    """
    Registra log no buffer e também no console
    """
    LOG_BUFFER.append(str(msg))
    print(msg)
    
    
# =============================================================
# VALIDAÇÃO DE VALORES
# =============================================================

def validar_campo(nome_campo, valor, valores_validos, default=None, linha_info=""):
    valor_u = str(valor).strip().upper()

    if valor_u not in valores_validos:
        if default is not None:
            log(
                f"[WARN] {linha_info} Campo {nome_campo} inválido: '{valor}'. "
                f"Usando '{default}'. Esperado: {sorted(valores_validos)}"
            )
            return default
        else:
            log(
                f"[ERROR] {linha_info} Campo {nome_campo} inválido: '{valor}'. "
                f"Esperado: {sorted(valores_validos)}"
            )
            return None

    return valor_u

def validar_softx(valor, linha_info=""):
    softx = str(valor).strip().upper()

    if softx not in SOFTX_SIGLA_MAP:
        log(
            f"[WARN] {linha_info} SOFTX desconhecido: '{valor}'. "
            f"Não encontrado no mapa de equivalência."
        )
        return softx, None

    return softx, SOFTX_SIGLA_MAP[softx]

def validar_cn(valor, linha_info=""):
    try:
        cn_int = int(valor)
    except (TypeError, ValueError):
        log(f"[ERROR] {linha_info} CN inválido (não numérico): '{valor}'")
        return None

    if cn_int not in VALID_CN:
        log(
            f"[ERROR] {linha_info}CN inválido: '{cn_int}'. "
            f"CN permitido conforme regra de rede."
        )
        return None

    return str(cn_int).zfill(2)



# =============================================================
# ROTAS FLASK (mantidas)
# =============================================================
@app.route('/',methods=['GET'])
def index(): return render_template('index.html')

@app.route("/gerar", methods=["POST"])
def gerar():
    f = request.files.get("arquivo")
    if not f or f.filename == "":
        flash("Selecione um arquivo.")
        return redirect(url_for("index"))

    if not allowed_file(f.filename):
        flash("Formato inválido.")
        return redirect(url_for("index"))

    df = load_dataframe(f)
    
    txt = gerar_txt(df)
    
    
    if txt is None:
            flash("Erros críticos encontrados. TXT não foi gerado. Verifique o log.")
            return redirect(url_for("index"))
    
    
    nome = (
        f"{safe_filename(df.iloc[0]['DOC'])}_"
        f"{safe_filename(df.iloc[0]['PST'])}_"
        f"{safe_filename(df.iloc[0]['TRAF'])}_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    )

    
    log("[OK] Script gerado com sucesso.")
    
    
    #SALVA COPIA DO TXT NO SERVIDOR
    path_txt = os.path.join(TXT_DIR, nome)
    with open(path_txt, "w", encoding="utf-8") as f:
        f.write(txt)

    return send_file(
        BytesIO(txt.encode("utf-8")),
        as_attachment=True,
        download_name=nome,
        mimetype="text/plain; charset=utf-8"
    )


@app.route('/modelo/<formato>')
def download_modelo(formato):
    if formato not in ('xlsx','csv'):
        flash('Formato inválido.'); return redirect(url_for('index'))
    fname='MODELO_PLANILHA_ROTAS.xlsx' if formato=='xlsx' else 'MODELO_PLANILHA_ROTAS.csv'
    return send_from_directory(MODELOS_DIR,fname,as_attachment=True)

@app.route("/gerar_planilha", methods=["POST"])
def gerar_planilha():
    f = request.files.get("arquivo")
    if not f or f.filename == "":
        flash("Selecione um arquivo.")
        return redirect(url_for("index"))

    if not allowed_file(f.filename):
        flash("Formato inválido.")
        return redirect(url_for("index"))

    df = load_dataframe(f)
    xlsx = gerar_planilha_ativacao(df)

    nome = (
        f"{safe_filename(df.iloc[0]['DOC'])}_"
        f"{safe_filename(df.iloc[0]['PST'])}_"
        f"{safe_filename(df.iloc[0]['TRAF'])}_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


    
    if not xlsx:
        flash("Erros críticos encontrados. Planilha não foi gerada. Verifique o log.")
        return redirect(url_for("index"))
    
    xlsx.seek(0)  # 🔑 garante ponteiro no início
    
    log("[OK] Planilha gerada com sucesso.")
    
    #SALVA COPIA DO XLSX NO SERVIDOR
    path_xlsx = os.path.join(XLSX_DIR, nome)
    with open(path_xlsx, "wb") as f:
        f.write(xlsx.getvalue())

    
    return send_file(
        xlsx,
        as_attachment=True,
        download_name=nome,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    
#===================================
# ROTA ENVIAR E-MAILS
#===================================

@app.route("/enviar_email", methods=["POST"])
def enviar_email():
    f = request.files.get("arquivo")
    ativado_por = request.form.get("ativado_por", "").strip()

    if not f or f.filename == "":
        flash("Selecione um arquivo.", "warning")
        return redirect(url_for("index"))

    if not allowed_file(f.filename):
        flash("Formato inválido.", "error")
        return redirect(url_for("index"))

    if not ativado_por:
        flash("Informe quem ativou a rota.", "warning")
        return redirect(url_for("index"))

    df = load_dataframe(f)

    # Gera a planilha completa (usada como base do HTML)
    xlsx_bytes = gerar_planilha_ativacao(df)

    if not xlsx_bytes:
        log("[ERROR] Não foi possível gerar a planilha para envio de e-mail.")
        flash(
            "Não foi possível enviar os e-mails porque a planilha contém erros. "
            "Verifique o log.",
            "error"
        )
        return redirect(url_for("index"))

    # Salva XLSX temporário para leitura das abas
    temp_dir = os.path.join(BASE_DIR, "temp")
    os.makedirs(temp_dir, exist_ok=True)

    xlsx_path = os.path.join(temp_dir, "ativacao.xlsx")
    with open(xlsx_path, "wb") as fp:
        fp.write(xlsx_bytes.getvalue())

    wb = load_workbook(xlsx_path)

    total_emails = len(df)
    emails_enviados = 0
    houve_erro_email = False

    log(f"[INFO] Iniciando envio de e-mails (total: {total_emails})...")

    # ✅ UM ÚNICO LOOP: GERA + ENVIA
    for idx, (_, r) in enumerate(df.iterrows(), start=1):

        tipo = str(r["TIPO DE ROTA"]).strip().upper()
        cn = str(int(r["CN"])).zfill(2)
        pti = str(r["PTI"]).strip()
        rrr = str(r["NOME DA ROTA"]).strip()
        rota_ebt = str(r.get("ROTA EBT", "")).strip()

        nome_aba = f"{tipo}_{cn}_{pti}"
        ws = wb[nome_aba]

        html_tabela = aba_para_html(ws)

        assunto = f'***Ativação de rota SIP - {rota_ebt}***'

        corpo_html = f"""
        <p><strong>Ativado por:</strong> {ativado_por}</p>
        <p>Segue os dados da rota:</p>
        {html_tabela}
        """

        contexto = f'Rota={rrr} | CN={cn} | TIPO={tipo} | PTI={pti}'

        log(f"[INFO] Enviando e-mail {idx} de {total_emails}... {contexto}")

        enviado = enviar_email_smtp_html(
            destinatarios="robinson.barreto@claro.com.br",
            assunto=assunto,
            html_body=corpo_html,
            contexto=contexto
        )

        if enviado:
            emails_enviados += 1
        else:
            houve_erro_email = True

    os.remove(xlsx_path)

    if emails_enviados > 0:
        if emails_enviados == 1:
            flash("1 e-mail enviado com sucesso.", "success")
        else:
            flash(f"{emails_enviados} e-mails enviados com sucesso.", "success")

        log("[OK] Envio de e-mail(s) concluído.")

    if houve_erro_email:
        flash(
            "Não foi possível enviar um ou mais e-mails "
            "(restrição de rede corporativa). Verifique o log.",
            "error"
        )

    return redirect(url_for("index"))

@app.route("/logs")
def get_logs():
    return {
        "logs": list(LOG_BUFFER)
    }
    
# limpar tela
@app.route("/limpar_logs", methods=["POST"])
def limpar_logs():
    LOG_BUFFER.clear()
    return "", 204


if __name__=='__main__':
    app.run(host='0.0.0.0',port=5050,debug=True)    
